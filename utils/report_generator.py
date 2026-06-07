import os
import datetime
from fpdf import FPDF, XPos, YPos


# ─── Colores INCES ───────────────────────────────────────────
INCES_BLUE  = (0,   64,  133)   # #004085
INCES_TEAL  = (0,  150,  166)   # #0096a6
WHITE       = (255, 255, 255)
LIGHT_GREY  = (245, 247, 250)
DARK_TEXT   = (33,  37,   41)
MID_GREY    = (108, 117,  125)


class INCESReport(FPDF):
    """
    Clase FPDF personalizada con cabecera y pie de página institucionales.
    """

    def __init__(self, title="Reporte INCES", subtitle=""):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.report_title    = title
        self.report_subtitle = subtitle
        self.generated_at    = datetime.datetime.now().strftime("%d/%m/%Y  %H:%M")

        # Fuente base (helvetica integrada, sin necesidad de archivos externos)
        self.set_auto_page_break(auto=True, margin=18)
        self.add_page()

    # ── Cabecera ─────────────────────────────────────────────
    def header(self):
        logo_path = os.path.join(os.path.dirname(__file__), "..", "Logo INCES.png")
        logo_path = os.path.normpath(logo_path)

        # Banda superior azul
        self.set_fill_color(*INCES_BLUE)
        self.rect(0, 0, self.w, 22, style="F")

        # Logo (si existe)
        if os.path.exists(logo_path):
            self.image(logo_path, x=6, y=2, h=18, keep_aspect_ratio=True)

        # Texto institucional
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(*WHITE)
        self.set_xy(0, 4)
        self.cell(self.w, 7, "INSTITUTO NACIONAL DE CAPACITACIÓN Y EDUCACIÓN SOCIALISTA", align="C")

        self.set_font("Helvetica", "", 8)
        self.set_xy(0, 11)
        self.cell(self.w, 5, "INCES  |  Sistema de Gestión de Censo e Inscripción", align="C")

        # Línea decorativa teal
        self.set_fill_color(*INCES_TEAL)
        self.rect(0, 22, self.w, 1.5, style="F")

        # Título del reporte
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*INCES_BLUE)
        self.set_xy(0, 25)
        self.cell(self.w, 7, self.report_title.upper(), align="C")

        if self.report_subtitle:
            self.set_font("Helvetica", "I", 9)
            self.set_text_color(*MID_GREY)
            self.set_xy(0, 32)
            self.cell(self.w, 5, self.report_subtitle, align="C")

        self.ln(22)   # Margen debajo de la cabecera

    # ── Pie de página ─────────────────────────────────────────
    def footer(self):
        self.set_y(-13)

        # Línea separadora
        self.set_draw_color(*INCES_TEAL)
        self.set_line_width(0.3)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(1)

        self.set_font("Helvetica", "I", 7)
        self.set_text_color(*MID_GREY)

        # Fecha a la izquierda
        self.cell(
            self.w / 3, 5,
            f"Generado: {self.generated_at}",
            align="L"
        )
        # Confidencial al centro
        self.cell(
            self.w / 3, 5,
            "Documento de uso interno - INCES",
            align="C"
        )
        # Número de página a la derecha
        self.cell(
            self.w / 3, 5,
            f"Pág. {self.page_no()} / {{nb}}",
            align="R",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )


# ─── Funciones de utilidad y generación ──────────────────────

def parse_fecha(fecha_str):
    """
    Analiza una cadena de fecha intentando múltiples formatos comunes de Google Sheets y SQLite.
    """
    if not fecha_str:
        return datetime.datetime.now()
    
    # Si ya es un objeto datetime o date
    if isinstance(fecha_str, (datetime.datetime, datetime.date)):
        return datetime.datetime.combine(fecha_str, datetime.time.min) if isinstance(fecha_str, datetime.date) else fecha_str

    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y"
    ]
    for fmt in formats:
        try:
            return datetime.datetime.strptime(fecha_str.strip(), fmt)
        except ValueError:
            continue
    try:
        # Intento manual para cadenas como "4/05/2026 22:20:25"
        parts = fecha_str.strip().split()
        if parts:
            date_part = parts[0]
            if "/" in date_part:
                d, m, y = map(int, date_part.split("/"))
                if y < 100:
                    y += 2000
                return datetime.datetime(y, m, d)
            elif "-" in date_part:
                y, m, d = map(int, date_part.split("-"))
                return datetime.datetime(y, m, d)
    except Exception:
        pass
    return datetime.datetime.now()


def generate_estudiantes_report(estudiantes: list, output_path: str | None = None, group_by_trimester: bool = True) -> str:
    """
    Genera un reporte PDF con la lista de estudiantes.

    Args:
        estudiantes: lista de sqlite3.Row o diccionarios.
        output_path: ruta donde guardar el PDF. Si es None, usa el directorio
                     'reports/' dentro de la raíz del proyecto.
        group_by_trimester: Si True, agrupa por trimestre; si False, lista general.

    Returns:
        Ruta absoluta del archivo PDF generado.
    """
    estudiantes = [dict(est) if not isinstance(est, dict) else est for est in estudiantes]

    if output_path is None:
        base_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
        reports_dir = os.path.join(base_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        timestamp  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = "General" if not group_by_trimester else "Trimestral"
        output_path = os.path.join(reports_dir, f"Reporte_Estudiantes_{suffix}_{timestamp}.pdf")

    total = len(estudiantes)
    today_str = datetime.datetime.now().strftime('%d/%m/%Y')
    subtitle = f"Total de participantes registrados: {total}  |  Fecha: {today_str}"

    title = "Listado General de Estudiantes" if not group_by_trimester else "Listado General de Estudiantes Censados por Trimestre"
    pdf = INCESReport(title=title, subtitle=subtitle)
    pdf.alias_nb_pages()

    # ─── Todas las columnas del formulario ─────────────────────
    col_widths = {
        "#":               7,
        "Cédula":         19,
        "Nombres":        20,
        "Apellidos":      20,
        "Género":          8,
        "Edad":            7,
        "Nivel Acad.":    18,
        "Discapacidad":   10,
        "Teléfono":       16,
        "Correo":         22,
        "Dirección":      22,
        "Entidad":        20,
        "Perfil":         22,
        "Estado":         15,
        "Fecha Censo":    16,
    }
    col_keys = list(col_widths.keys())
    row_h = 6

    def draw_table_header():
        pdf.set_fill_color(*INCES_TEAL)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 6.5)
        for label in col_keys:
            pdf.cell(col_widths[label], row_h + 1, label, border=0, align="C", fill=True,
                     new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.ln(row_h + 1)

    # Tratar None / texto largo
    def safe(val, max_chars=30):
        return str(val or "")[:max_chars]

    estado_colors = {
        "INSCRITO":  (0, 128, 0),
        "CENSADO":   (204, 140, 0),
        "RETIRADO":  (180, 0, 0),
        "CULMINADO": (0, 100, 180),
    }

    def write_row(est, row_idx):
        fill_color = row_colors[row_idx % 2]
        pdf.set_fill_color(*fill_color)
        pdf.set_font("Helvetica", "", 6.5)
        pdf.set_text_color(*DARK_TEXT)

        estado = est.get("estado_inscripcion") or "CENSADO"
        perfil = est.get("perfil_nombre") or "Sin asignar"

        discap = est.get("posee_discapacidad")
        discap_str = "Sí" if discap and str(discap).lower() in ("1", "sí", "si", "yes", "true") else "No"

        cells = [
            (str(row_idx + 1), "#", "C"),
            (safe(est.get("cedula"), 12), "Cédula", "C"),
            (safe(est.get("nombres"), 14), "Nombres", "L"),
            (safe(est.get("apellidos"), 14), "Apellidos", "L"),
            (safe(est.get("genero"), 6), "Género", "C"),
            (str(est.get("edad") or ""), "Edad", "C"),
            (safe(est.get("nivel_academico"), 14), "Nivel Acad.", "L"),
            (discap_str, "Discapacidad", "C"),
            (safe(est.get("telefono"), 12), "Teléfono", "C"),
            (safe(est.get("correo"), 18), "Correo", "L"),
            (safe(est.get("direccion"), 18), "Dirección", "L"),
            (safe(est.get("entidad"), 15), "Entidad", "L"),
            (safe(perfil, 16), "Perfil", "L"),
        ]

        for text, key, align in cells:
            pdf.cell(col_widths[key], row_h, f" {text}", border="B", align=align, fill=True,
                     new_x=XPos.RIGHT, new_y=YPos.TOP)

        # Celda de estado con color propio
        pdf.set_text_color(*estado_colors.get(estado, DARK_TEXT))
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.cell(col_widths["Estado"], row_h, estado, border="B", align="C", fill=True,
                 new_x=XPos.RIGHT, new_y=YPos.TOP)

        pdf.set_text_color(*DARK_TEXT)
        pdf.set_font("Helvetica", "", 6.5)
        f_censo = safe(est.get("fecha_censo"), 12)
        pdf.cell(col_widths["Fecha Censo"], row_h, f" {f_censo}", border="B", align="C", fill=True,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    global_row_index = 0
    row_colors = [WHITE, LIGHT_GREY]

    if group_by_trimester:
        # ─── Agrupado por trimestre ────────────────────────────
        groups = {}
        for est in estudiantes:
            dt = parse_fecha(est.get("fecha_censo"))
            quarter = (dt.month - 1) // 3 + 1
            key = (dt.year, quarter)
            if key not in groups:
                groups[key] = []
            groups[key].append(est)

        sorted_keys = sorted(groups.keys(), key=lambda x: (x[0], x[1]))

        for idx, (year, quarter) in enumerate(sorted_keys):
            quarter_est = groups[(year, quarter)]

            if pdf.get_y() + 28 > pdf.page_break_trigger:
                pdf.add_page()
            else:
                if idx > 0 and pdf.get_y() > 35:
                    pdf.ln(4)

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*INCES_BLUE)
            pdf.cell(0, 6, f"TRIMESTRE {quarter} - AÑO {year}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(1)
            draw_table_header()

            for est in quarter_est:
                if pdf.get_y() + row_h > pdf.page_break_trigger:
                    pdf.add_page()
                    pdf.set_font("Helvetica", "B", 10)
                    pdf.set_text_color(*INCES_BLUE)
                    pdf.cell(0, 6, f"TRIMESTRE {quarter} - AÑO {year} (Continuación)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.ln(1)
                    draw_table_header()

                write_row(est, global_row_index)
                global_row_index += 1
    else:
        # ─── General (sin agrupar) ─────────────────────────────
        draw_table_header()
        for est in estudiantes:
            if pdf.get_y() + row_h > pdf.page_break_trigger:
                pdf.add_page()
                draw_table_header()
            write_row(est, global_row_index)
            global_row_index += 1

    # ─── Sección de resumen estadístico ──────────────────────
    pdf.ln(6)
    pdf.set_fill_color(*INCES_BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 7, "  RESUMEN ESTADÍSTICO", fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    from collections import Counter
    estados_count = Counter(est.get("estado_inscripcion") or "CENSADO" for est in estudiantes)
    genero_count  = Counter(est.get("genero") or "No especificado" for est in estudiantes)

    pdf.set_fill_color(*LIGHT_GREY)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.ln(2)

    pdf.cell(0, 5, f"  Total de estudiantes: {total}", fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for estado, cnt in estados_count.most_common():
        pdf.cell(0, 5, f"      {estado}: {cnt}", fill=True,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(2)
    pdf.cell(0, 5, "  Distribución por Género:", fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for genero, cnt in genero_count.most_common():
        pdf.cell(0, 5, f"      {genero}: {cnt}", fill=True,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.output(output_path)
    return os.path.abspath(output_path)


def generate_estudiantes_xlsx_report(estudiantes: list, output_path: str | None = None, group_by_trimester: bool = True) -> str:
    """
    Genera un reporte Excel (.xlsx) con la lista de estudiantes.

    Args:
        estudiantes: lista de sqlite3.Row o diccionarios.
        output_path: ruta donde guardar el Excel. Si es None, usa el directorio
                     'reports/' dentro de la raíz del proyecto.
        group_by_trimester: Si True, agrupa por trimestre; si False, lista general.

    Returns:
        Ruta absoluta del archivo Excel generado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    estudiantes = [dict(est) if not isinstance(est, dict) else est for est in estudiantes]

    if output_path is None:
        base_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
        reports_dir = os.path.join(base_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        timestamp  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = "General" if not group_by_trimester else "Trimestral"
        output_path = os.path.join(reports_dir, f"Reporte_Estudiantes_{suffix}_{timestamp}.xlsx")

    # ─── Configuración de Estilos ─────────────────────────────
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, color="FFFFFF")
    font_main_title = Font(name="Segoe UI", size=12, bold=True, color="004085")
    font_meta = Font(name="Segoe UI", size=9, italic=True, color="555555")
    font_sec_title = Font(name="Segoe UI", size=11, bold=True, color="004085")
    font_tbl_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="000000")
    font_bold_data = Font(name="Segoe UI", size=10, bold=True, color="000000")

    fill_blue = PatternFill(start_color="002D5A", end_color="002D5A", fill_type="solid")
    fill_teal = PatternFill(start_color="0096A6", end_color="0096A6", fill_type="solid")
    fill_sec_bg = PatternFill(start_color="E6F2F5", end_color="E6F2F5", fill_type="solid")
    fill_zebra = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    estado_colors = {
        "INSCRITO":  "008000",
        "CENSADO":   "CC8C00",
        "RETIRADO":  "B40000",
        "RECHAZADO": "B40000",
        "CULMINADO": "0064B4"
    }

    # ─── Todas las columnas ───────────────────────────────────
    headers = [
        "#", "Cédula", "Nombres", "Apellidos", "Género", "Edad",
        "Nivel Académico", "Discapacidad", "Teléfono", "Correo",
        "Dirección", "Nombre de Entidad", "Perfil", "Estado", "Fecha Censo"
    ]
    num_cols = len(headers)
    last_col_letter = get_column_letter(num_cols)

    # Agrupar si aplica
    if group_by_trimester:
        groups = {}
        for est in estudiantes:
            dt = parse_fecha(est.get("fecha_censo"))
            quarter = (dt.month - 1) // 3 + 1
            key = (dt.year, quarter)
            if key not in groups:
                groups[key] = []
            groups[key].append(est)
        sorted_keys = sorted(groups.keys(), key=lambda x: (x[0], x[1]))
    else:
        sorted_keys = [(0, 0)]
        groups = {(0, 0): estudiantes}

    # ─── Crear libro ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    ws.views.sheetView[0].showGridLines = True

    # Banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "INSTITUTO NACIONAL DE CAPACITACIÓN Y EDUCACIÓN SOCIALISTA"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_blue
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 25

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "INCES  |  Sistema de Gestión de Censo e Inscripción"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = fill_blue
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 4
    for col in range(1, num_cols + 1):
        ws.cell(row=3, column=col).fill = fill_teal

    title_text = "LISTADO GENERAL DE ESTUDIANTES" if not group_by_trimester else "LISTADO GENERAL DE ESTUDIANTES CENSADOS POR TRIMESTRE"
    ws.merge_cells(f"A4:{last_col_letter}4")
    ws["A4"] = title_text
    ws["A4"].font = font_main_title
    ws["A4"].alignment = align_center
    ws.row_dimensions[4].height = 25

    total_est = len(estudiantes)
    generated_str = datetime.datetime.now().strftime("%d/%m/%Y  %H:%M")
    ws.merge_cells(f"A5:{last_col_letter}5")
    ws["A5"] = f"Total Participantes: {total_est}  |  Generado: {generated_str}  |  Documento de uso interno - INCES"
    ws["A5"].font = font_meta
    ws["A5"].alignment = align_center
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 10

    # ─── Datos ────────────────────────────────────────────────
    curr_row = 7
    global_idx = 1

    for (year, quarter) in sorted_keys:
        q_est = groups[(year, quarter)]

        if group_by_trimester:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=num_cols)
            sec_cell = ws.cell(row=curr_row, column=1)
            sec_cell.value = f"TRIMESTRE {quarter} - AÑO {year}"
            sec_cell.font = font_sec_title
            sec_cell.fill = fill_sec_bg
            sec_cell.alignment = align_left
            ws.row_dimensions[curr_row].height = 24
            for col in range(1, num_cols + 1):
                ws.cell(row=curr_row, column=col).fill = fill_sec_bg
                ws.cell(row=curr_row, column=col).border = border_thin
            curr_row += 1

        # Encabezados de columna
        ws.row_dimensions[curr_row].height = 20
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.value = header_text
            cell.font = font_tbl_header
            cell.fill = fill_teal
            cell.alignment = align_center
            cell.border = border_thin
        curr_row += 1

        row_colors = [fill_white, fill_zebra]

        for est in q_est:
            ws.row_dimensions[curr_row].height = 18
            fill_color = row_colors[global_idx % 2]

            estado = est.get("estado_inscripcion") or "CENSADO"
            perfil = est.get("perfil_nombre") or "Sin asignar"
            discap = est.get("posee_discapacidad")
            discap_str = "Sí" if discap and str(discap).lower() in ("1", "sí", "si", "yes", "true") else "No"

            data = [
                global_idx,
                str(est.get("cedula") or "").strip(),
                est.get("nombres") or "",
                est.get("apellidos") or "",
                est.get("genero") or "",
                est.get("edad") or "",
                est.get("nivel_academico") or "",
                discap_str,
                est.get("telefono") or "N/A",
                est.get("correo") or "",
                est.get("direccion") or "",
                est.get("entidad") or "",
                perfil,
                estado,
                est.get("fecha_censo") or "",
            ]

            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.border = border_thin
                cell.fill = fill_color
                if col_idx == 7 or col_idx == 13:  # Nivel Académico, Estado
                    cell.font = font_data
                else:
                    cell.font = font_data
                if col_idx == 14:  # Estado
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color=estado_colors.get(estado, "000000"))
                cell.alignment = align_center if col_idx in (1, 2, 5, 6, 8, 9, 14, 15) else align_left

            global_idx += 1
            curr_row += 1

        if group_by_trimester:
            ws.row_dimensions[curr_row].height = 12
            curr_row += 1

    # ─── Resumen Estadístico ──────────────────────────────────
    ws.row_dimensions[curr_row].height = 20
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
    res_cell = ws.cell(row=curr_row, column=1)
    res_cell.value = "RESUMEN ESTADÍSTICO"
    res_cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    res_cell.fill = fill_blue
    res_cell.alignment = align_center
    for col in range(1, 5):
        ws.cell(row=curr_row, column=col).fill = fill_blue
        ws.cell(row=curr_row, column=col).border = border_thin
    curr_row += 1

    from collections import Counter
    estados_count = Counter(est.get("estado_inscripcion") or "CENSADO" for est in estudiantes)
    genero_count  = Counter(est.get("genero") or "No especificado" for est in estudiantes)

    ws.row_dimensions[curr_row].height = 18
    ws.cell(row=curr_row, column=1, value="Métrica / Categoría").font = font_tbl_header
    ws.cell(row=curr_row, column=1).fill = fill_teal
    ws.cell(row=curr_row, column=1).alignment = align_left
    ws.cell(row=curr_row, column=1).border = border_thin
    ws.cell(row=curr_row, column=2, value="Cantidad").font = font_tbl_header
    ws.cell(row=curr_row, column=2).fill = fill_teal
    ws.cell(row=curr_row, column=2).alignment = align_center
    ws.cell(row=curr_row, column=2).border = border_thin
    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
    c_aux = ws.cell(row=curr_row, column=3, value="Porcentaje")
    c_aux.font = font_tbl_header
    c_aux.fill = fill_teal
    c_aux.alignment = align_center
    for col in range(3, 5):
        ws.cell(row=curr_row, column=col).fill = fill_teal
        ws.cell(row=curr_row, column=col).border = border_thin
    curr_row += 1

    ws.row_dimensions[curr_row].height = 18
    ws.cell(row=curr_row, column=1, value="Total Estudiantes").font = font_bold_data
    ws.cell(row=curr_row, column=1).border = border_thin
    ws.cell(row=curr_row, column=2, value=total_est).font = font_bold_data
    ws.cell(row=curr_row, column=2).alignment = align_center
    ws.cell(row=curr_row, column=2).border = border_thin
    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
    ws.cell(row=curr_row, column=3, value="100.0%").font = font_bold_data
    ws.cell(row=curr_row, column=3).alignment = align_center
    for col in range(3, 5):
        ws.cell(row=curr_row, column=col).border = border_thin
    curr_row += 1

    for estado, cnt in estados_count.most_common():
        pct = (cnt / total_est * 100) if total_est > 0 else 0
        ws.row_dimensions[curr_row].height = 18
        ws.cell(row=curr_row, column=1, value=f"  Estado: {estado}").font = font_data
        ws.cell(row=curr_row, column=1).border = border_thin
        ws.cell(row=curr_row, column=2, value=cnt).font = font_data
        ws.cell(row=curr_row, column=2).alignment = align_center
        ws.cell(row=curr_row, column=2).border = border_thin
        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
        ws.cell(row=curr_row, column=3, value=f"{pct:.1f}%").font = font_data
        ws.cell(row=curr_row, column=3).alignment = align_center
        for col in range(3, 5):
            ws.cell(row=curr_row, column=col).border = border_thin
        curr_row += 1

    curr_row += 1
    ws.row_dimensions[curr_row].height = 18
    ws.cell(row=curr_row, column=1, value="Distribución por Género").font = font_bold_data
    ws.cell(row=curr_row, column=1).border = border_thin
    ws.cell(row=curr_row, column=2, value="").border = border_thin
    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
    for col in range(3, 5):
        ws.cell(row=curr_row, column=col).border = border_thin
    curr_row += 1

    for genero, cnt in genero_count.most_common():
        pct = (cnt / total_est * 100) if total_est > 0 else 0
        ws.row_dimensions[curr_row].height = 18
        ws.cell(row=curr_row, column=1, value=f"  {genero}").font = font_data
        ws.cell(row=curr_row, column=1).border = border_thin
        ws.cell(row=curr_row, column=2, value=cnt).font = font_data
        ws.cell(row=curr_row, column=2).alignment = align_center
        ws.cell(row=curr_row, column=2).border = border_thin
        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
        ws.cell(row=curr_row, column=3, value=f"{pct:.1f}%").font = font_data
        ws.cell(row=curr_row, column=3).alignment = align_center
        for col in range(3, 5):
            ws.cell(row=curr_row, column=col).border = border_thin
        curr_row += 1

    # ─── Auto-ajuste de Anchos de Columna ──────────────────────
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 4, 5]:
                continue
            if cell.value and isinstance(cell.value, str) and ("TRIMESTRE" in cell.value or "RESUMEN" in cell.value or "INSTITUTO" in cell.value):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_path)
    return os.path.abspath(output_path)
